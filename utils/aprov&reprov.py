import pandas as pd
from openpyxl import load_workbook


CAMINHO_ARQUIVO = r"QUADRO AUTOMATIZADO.xlsx"
URL_SHEET = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRyWd4XSMVqWzpui9dzYqK65AZJZk49wobW5ekP10VIZ9VD4TavDxmojf_wtkG5TZRnKNoeaqcK8RfU/pub?gid=1321843174&single=true&output=csv"

def aprovar_transferencia(nome, nova_loja):
    try:
        df = pd.read_csv(URL_SHEET)
        wb = load_workbook(CAMINHO_ARQUIVO)
        ws = wb['QUADRO']

        encontrado = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Coluna C
            celula = row[0]
            if celula.value and str(celula.value).strip().upper() == nome.strip().upper():
                linha = celula.row
                ws[f"E{linha}"] = nova_loja # Coluna da loja
                encontrado = True
                break

        wb.save(CAMINHO_ARQUIVO)
        wb.close()

        if encontrado:
            print(f"✅ {nome} transferido para loja {nova_loja} (APROVADO)")
        else:
            print(f"⚠️ Nome '{nome}' não encontrado na planilha.")

    except Exception as e:
        print(f"Erro ao aprovar transferência: {e}")


def negar_transferencia(nome):
    try:
        wb = load_workbook(CAMINHO_ARQUIVO)
        ws = wb['QUADRO']

        encontrado = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Coluna C
            celula = row[0]
            if celula.value and str(celula.value).strip().upper() == nome.strip().upper():
                linha = celula.row # Coluna de status
                encontrado = True
                break

        wb.save(CAMINHO_ARQUIVO)
        wb.close()

        if encontrado:
            print(f"❌ Transferência de {nome} foi NEGADA")
        else:
            print(f"⚠️ Nome '{nome}' não encontrado na planilha.")

    except Exception as e:
        print(f"Erro ao negar transferência: {e}")